# exporter.py
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from utils import ts, extract_video_id  # adjust import path if in a package

class ExcelChapterExporter:
    """
    Writes chapters + per-chapter concepts to an .xlsx file.
    - One row per chapter (title, start/end with hyperlinks, summary)
    - Then one row per concept under that chapter
    - Each mention timestamp is a separate hyperlinked cell
    """

    def __init__(self, video_url: str, out_path: str = "chapters.xlsx"):
        self.video_url = self._canonical_watch_url(video_url)
        self.out_path = out_path

    # ---- public ----
    def export(self, chapters_result: Dict[str, Any]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chapters+Concepts"

        # Compute max number of mentions across all concepts to size columns
        max_mentions = self._max_mentions(chapters_result["chapters"])

        # Header
        headers = ["Chapter #", "Title", "Start", "End", "Source", "Summary", "Concept"]
        headers += [f"Mention {i+1}" for i in range(max_mentions)]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)

        source_label = "YouTube (yt-dlp)" if chapters_result["source"] == "official" else "LLM (GPT-4o)"

        # Rows
        row_idx = 2
        for i, ch in enumerate(chapters_result["chapters"], start=1):
            # Chapter row
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=ch["title"])

            # Start/end as hyperlinked [MM:SS]
            start_lbl = ts(ch["start"])
            end_lbl = ts(ch["end"])
            start_cell = ws.cell(row=row_idx, column=3, value=start_lbl)
            end_cell = ws.cell(row=row_idx, column=4, value=end_lbl)
            start_cell.hyperlink = self._yt_link(ch["start"])
            end_cell.hyperlink = self._yt_link(ch["end"])

            # Source + Summary
            ws.cell(row=row_idx, column=5, value=source_label)
            summary_cell = ws.cell(row=row_idx, column=6, value=ch.get("summary", ""))
            summary_cell.alignment = Alignment(wrap_text=True)

            # Advance to concept rows
            row_idx += 1

            # Concept rows (under the chapter)
            for concept in ch.get("concepts", []):
                ws.cell(row=row_idx, column=1, value=i)  # keep chapter # for filtering
                ws.cell(row=row_idx, column=7, value=concept.get("name", ""))

                mentions: List[str] = concept.get("mentions", [])
                for m_idx, stamp in enumerate(mentions[:max_mentions]):
                    # stamp looks like "[MM:SS]"; parse to seconds for link
                    try:
                        mm, ss = stamp.strip("[]").split(":")
                        seconds = int(mm) * 60 + int(ss)
                    except Exception:
                        seconds = 0
                    col = 8 + m_idx  # first mention col
                    cell = ws.cell(row=row_idx, column=col, value=stamp)
                    cell.hyperlink = self._yt_link(seconds)

                row_idx += 1

            # blank spacer row between chapters
            row_idx += 1

        # Column widths
        widths = {
            1: 11,   # Chapter #
            2: 40,   # Title
            3: 10,   # Start
            4: 10,   # End
            5: 16,   # Source
            6: 80,   # Summary
            7: 36,   # Concept
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Mentions columns sized modestly
        for col_idx in range(8, 8 + max_mentions):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        wb.save(self.out_path)
        return self.out_path

    # ---- helpers ----
    def _max_mentions(self, chapters: List[Dict[str, Any]]) -> int:
        m = 1
        for ch in chapters:
            for c in ch.get("concepts", []):
                m = max(m, len(c.get("mentions", [])))
        return m

    def _canonical_watch_url(self, video_url: str) -> str:
        # normalize to https://www.youtube.com/watch?v=<id> so &t= works cleanly
        vid = extract_video_id(video_url)
        return f"https://www.youtube.com/watch?v={vid}"

    def _yt_link(self, seconds: float) -> str:
        secs = int(seconds)
        return f"{self.video_url}&t={secs}s"
